"""File processing service — extract text from various file formats."""

import csv
import io
import logging
from pathlib import Path

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {
    ".pdf": "application/pdf",
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".bmp": "image/bmp",
    ".tiff": "image/tiff",
    ".md": "text/markdown",
    ".txt": "text/plain",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".csv": "text/csv",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
}


class FileService:
    """Service for extracting text from various file formats."""

    def get_file_type(self, filename: str) -> str | None:
        """Get the file type from the filename extension."""
        ext = Path(filename).suffix.lower()
        return ext if ext in SUPPORTED_EXTENSIONS else None

    def is_supported(self, filename: str) -> bool:
        """Check if the file type is supported."""
        return self.get_file_type(filename) is not None

    async def extract_text(self, file_content: bytes, filename: str) -> str:
        """Extract text from a file based on its type."""
        file_type = self.get_file_type(filename)
        if file_type is None:
            raise ValueError(f"Unsupported file type: {filename}")

        extractors = {
            ".pdf": self._extract_pdf,
            ".png": self._extract_image,
            ".jpg": self._extract_image,
            ".jpeg": self._extract_image,
            ".bmp": self._extract_image,
            ".tiff": self._extract_image,
            ".md": self._extract_text,
            ".txt": self._extract_text,
            ".xlsx": self._extract_xlsx,
            ".csv": self._extract_csv,
            ".docx": self._extract_docx,
        }

        extractor = extractors.get(file_type)
        if extractor is None:
            raise ValueError(f"No extractor for file type: {file_type}")

        return await extractor(file_content, filename)

    async def _extract_pdf(self, content: bytes, filename: str) -> str:
        """Extract text from a PDF file using PyMuPDF."""
        import fitz  # pymupdf

        text_parts = []
        with fitz.open(stream=content, filetype="pdf") as doc:
            for page in doc:
                text_parts.append(page.get_text())
        return "\n\n".join(text_parts).strip()

    async def _extract_image(self, content: bytes, filename: str) -> str:
        """Extract text from an image using Tesseract OCR."""
        import pytesseract
        from PIL import Image

        image = Image.open(io.BytesIO(content))
        text = pytesseract.image_to_string(image, lang="eng+chi_sim+deu+fra")
        return text.strip()

    async def _extract_text(self, content: bytes, filename: str) -> str:
        """Extract text from plain text / markdown files."""
        return content.decode("utf-8", errors="replace").strip()

    async def _extract_xlsx(self, content: bytes, filename: str) -> str:
        """Extract text from an Excel file."""
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True)
        text_parts = []
        for sheet in wb:
            text_parts.append(f"## Sheet: {sheet.title}\n")
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                text_parts.append(" | ".join(cells))
        return "\n".join(text_parts).strip()

    async def _extract_csv(self, content: bytes, filename: str) -> str:
        """Extract text from a CSV file."""
        text = content.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = [" | ".join(row) for row in reader]
        return "\n".join(rows).strip()

    async def _extract_docx(self, content: bytes, filename: str) -> str:
        """Extract text from a DOCX file."""
        try:
            from docx import Document

            doc = Document(io.BytesIO(content))
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n\n".join(paragraphs).strip()
        except ImportError:
            logger.warning("python-docx not installed, skipping DOCX extraction")
            return "[DOCX extraction requires python-docx package]"


# Singleton instance
file_service = FileService()
